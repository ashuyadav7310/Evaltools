# src/output_generator.py [document]

import os
from datetime import datetime
from typing import Dict, List, Any
import json
import time
from openai import OpenAI
import re
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
from src.format_evidence import extract_format_evidence


OUTPUT_DIR = "./output"
os.makedirs(OUTPUT_DIR, exist_ok=True)


def _safe_dict(rec):
    """Guarantee rec is a dict."""
    return rec if isinstance(rec, dict) else {}


def _safe_list(x):
    """Force value into list."""
    if isinstance(x, list):
        return x
    if x:
        return [x]
    return []


def collect_weak_rubric_evidence(rubric_rows: list[dict]) -> list[dict]:
    evidence = []

    for row in rubric_rows:
        score = row["Score"]
        weight = row["Weight"]
        ratio = score / weight if weight else 0

        if ratio < 0.6:
            severity = "major"
        elif ratio < 0.75:
            severity = "moderate"
        else:
            continue

        evidence.append({
            "rubric": row["Rubric Name"],
            "feedback": row["Feedback"],
            "score": score,
            "weight": weight,
            "type": row["Type"],
            "severity": severity,
        })

    return evidence


def collect_strong_rubric_evidence(rubric_rows: list[dict]) -> list[dict]:
    evidence = []

    for row in rubric_rows:
        score = row["Score"]
        weight = row["Weight"]
        ratio = score / weight if weight else 0

        if ratio < 0.75:
            continue

        evidence.append({
            "rubric": row["Rubric Name"],
            "feedback": row["Feedback"],
            "score": score,
            "weight": weight,
            "type": row["Type"],
            "coverage": round(ratio, 2),
        })

    evidence.sort(key=lambda item: (item.get("coverage", 0), item.get("score", 0)), reverse=True)
    return evidence


client = None


def _get_openai_client():
    global client
    if client is not None:
        return client

    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        return None

    client = OpenAI(api_key=api_key)
    return client


def _sanitize_remark(text: str) -> str:
    if not text:
        return ""

    banned_phrases = [
        "would you like",
        "insert grade",
        "here is",
        "sample",
        "example",
        "let me know",
        "i can",
        "do you want",
    ]

    lowered = text.lower()
    if any(p in lowered for p in banned_phrases):
        return ""

    if len(text.split()) < 12:
        return ""

    cleaned = text.strip()
    cleaned = re.sub(r"\bmarks?\s+were\s+slightly\s+reduced\b", "scores reflect criterion performance", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"\bslightly\s+reduced\b", "lowered", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"\bminor\s+deduction[s]?\b", "deductions", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"\bmarks\s+are\s+reduced\b", "scores reflect criterion performance", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"\brubrics?\b", "criteria", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"\bmarks?\b", "scores", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    return cleaned


def _extract_feedback_cues(feedback: str) -> dict[str, str]:
    if not feedback:
        return {}

    labels = ("Strength", "Improve", "Rewrite tip", "Etiquette")
    out: dict[str, str] = {}

    for label in labels:
        match = re.search(
            rf"{label}\s*:\s*(.*?)(?=\s+(?:Strength|Improve|Rewrite tip|Etiquette)\s*:|$)",
            feedback,
            flags=re.IGNORECASE | re.DOTALL,
        )
        if match:
            out[label.lower().replace(" ", "_")] = match.group(1).strip()

    base = re.sub(
        r"\s*(?:Strength|Improve|Rewrite tip|Etiquette)\s*:.*$",
        "",
        feedback,
        flags=re.IGNORECASE | re.DOTALL,
    ).strip()
    if base:
        out["base"] = base

    return out


def _clean_feedback_fragment(text: str, limit: int = 180) -> str:
    if not text:
        return ""

    text = re.sub(r"\s+", " ", str(text)).strip()
    text = re.sub(r"^Avg sentence length=.*?Contractions=\d+\.\s*", "", text, flags=re.IGNORECASE)
    text = re.sub(r"^Keypoint coverage=.*?missed=\d+\.\s*", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\bImage relevance contributed positively.*$", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\bDiagram correctness supported clarity.*$", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\s+", " ", text).strip(" .;,-")

    if len(text) > limit:
        text = text[:limit].rsplit(" ", 1)[0].rstrip(" ,;:-") + "..."

    return text


def _serialize_evidence_for_prompt(evidence: list[dict], limit: int) -> list[dict]:
    payload = []
    for item in evidence[:limit]:
        cues = _extract_feedback_cues(str(item.get("feedback", "")))
        payload.append({
            "rubric": item.get("rubric", ""),
            "type": item.get("type", ""),
            "score": item.get("score", 0),
            "weight": item.get("weight", 0),
            "coverage": item.get("coverage"),
            "severity": item.get("severity"),
            "feedback": _clean_feedback_fragment(cues.get("base", "")),
            "strength_note": _clean_feedback_fragment(cues.get("strength", "")),
            "improvement_note": _clean_feedback_fragment(cues.get("improve", "")),
            "rewrite_tip": _clean_feedback_fragment(cues.get("rewrite_tip", "")),
            "etiquette": _clean_feedback_fragment(cues.get("etiquette", "")),
        })
    return payload


def _infer_submission_mode(rubric_rows: list[dict], answer_text: str) -> str:
    rubric_types = {
        str(row.get("Type", "")).strip().lower()
        for row in rubric_rows
        if row.get("Type")
    }
    if "email" in rubric_types or "email_format" in rubric_types:
        return "email"

    text = (answer_text or "").lower()
    if re.search(r"\bsubject\s*[:\-]", text) and re.search(r"\b(regards|sincerely|thank you|best regards)\b", text):
        return "email"

    return "document"


def _detect_rubric_category(rubric_rows: list[dict]) -> str:
    rubric_types = {
        str(row.get("Type", "")).strip().lower()
        for row in rubric_rows
        if row.get("Type")
    }

    if rubric_types & {"content", "format", "style"}:
        return "content"

    if rubric_types & {"email", "email_format"}:
        return "email"

    return "content"


def _extract_email_strength_and_improvement(
    strong_evidence: list[dict] | None,
    weak_evidence: list[dict] | None,
) -> tuple[str, str]:
    strong_evidence = strong_evidence or []
    weak_evidence = weak_evidence or []

    def _pick_feedback_fragment(feedback: str, label: str) -> str:
        if not feedback:
            return ""
        match = re.search(
            rf"{label}\s*:\s*(.*?)(?=\s+(?:Strength|Improve|Rewrite tip|Etiquette)\s*:|$)",
            feedback,
            flags=re.IGNORECASE | re.DOTALL,
        )
        if not match:
            return ""
        return _clean_feedback_fragment(match.group(1).strip(), limit=220)

    strength = ""
    for item in strong_evidence:
        strength = _pick_feedback_fragment(str(item.get("feedback", "")), "Strength")
        if strength:
            break

    if not strength:
        for item in weak_evidence:
            strength = _pick_feedback_fragment(str(item.get("feedback", "")), "Strength")
            if strength:
                break

    if not strength:
        if strong_evidence:
            strength = f"{strong_evidence[0].get('rubric', 'A key area')} is comparatively strong."
        else:
            strength = "Core message communication is clear."

    improvement = ""
    for item in weak_evidence:
        improvement = _pick_feedback_fragment(str(item.get("feedback", "")), "Improve")
        if improvement:
            break

    if not improvement:
        if weak_evidence:
            improvement = f"{weak_evidence[0].get('rubric', 'A key area')} needs better precision and structure."
        else:
            improvement = "Overall consistency and professional polish can be improved."

    return strength, improvement


def _build_personalized_remark_from_evidence(
    student_name: str,
    submission_mode: str,
    strong_evidence: list[dict] | None,
    weak_evidence: list[dict] | None,
    total_marks: float | None = None,
    max_marks: float | None = None,
) -> str:
    strong_evidence = strong_evidence or []
    weak_evidence = weak_evidence or []

    strong = strong_evidence[0] if strong_evidence else None
    weak = weak_evidence[0] if weak_evidence else None

    score_text = ""
    if total_marks is not None and max_marks:
        score_text = f" ({round(total_marks, 2)}/{round(max_marks, 2)})"

    if submission_mode == "email":
        if strong and weak:
            return (
                f"{student_name}{score_text} handles {strong.get('rubric', 'the strongest criterion')} well, "
                f"but loses ground in {weak.get('rubric', 'the weakest criterion')}. "
                f"The response shows the right business intent, yet the email is not equally strong in structure, tone, or execution."
            )
        if strong:
            return (
                f"{student_name}{score_text} delivers a credible email, especially in "
                f"{strong.get('rubric', 'the strongest criterion')}, though the response is not fully polished across all criteria."
            )
        return (
            f"{student_name}{score_text} addresses the email task, but the response remains uneven in professional structure and execution."
        )

    if strong and weak:
        return (
            f"{student_name}{score_text} is strongest in {strong.get('rubric', 'the strongest criterion')}, "
            f"while {weak.get('rubric', 'the weakest criterion')} limits the overall result. "
            f"The submission shows clear understanding of the case, but the weaker sections are not developed with the same depth, precision, or organization."
        )
    if strong:
        return (
            f"{student_name}{score_text} produces a strong submission, especially in "
            f"{strong.get('rubric', 'the strongest criterion')}, although the overall document is not equally sharp in every section."
        )
    if weak:
        return (
            f"{student_name}{score_text} covers the assignment, but {weak.get('rubric', 'one major criterion')} remains underdeveloped. "
            f"The response needs stronger depth and tighter structure to become consistently convincing."
        )
    return (
        f"{student_name}{score_text} addresses the task, but the submission still reads as broad rather than sharply evidenced."
    )


def _build_remark_prompt(
    student_name: str,
    submission_mode: str,
    band: str,
    total_marks: float | None,
    max_marks: float | None,
    strong_payload: list[dict],
    weak_payload: list[dict],
    doc_meta: dict[str, Any] | None,
) -> tuple[str, str]:
    doc_meta = doc_meta or {}
    paragraph_blocks = doc_meta.get("paragraph_blocks", 0)
    tables = len(doc_meta.get("tables", []) or [])
    images = doc_meta.get("images_detected", 0)

    if submission_mode == "email":
        system_prompt = (
            "You write sharp evaluator remarks for professional email submissions. "
            "Use only the supplied evidence. Avoid generic praise, filler, and template language."
        )
        user_prompt = f"""
TASK:
Write one compact evaluator remark for an email-writing submission.

STUDENT:
{student_name}

SCORE:
{total_marks if total_marks is not None else "N/A"} / {max_marks if max_marks is not None else "N/A"}

PERFORMANCE BAND:
{band}

TOP STRENGTHS:
{json.dumps(strong_payload, indent=2)}

TOP WEAKNESSES:
{json.dumps(weak_payload, indent=2)}

INSTRUCTIONS:
- Write exactly 2 or 3 sentences.
- Mention the student name once.
- Base every sentence on the supplied evidence.
- Focus on email quality: tone, structure, action clarity, professionalism, greeting, sign-off, and concision.
- Mention the strongest criterion by name.
- Mention the weakest criterion by name if one exists.
- If etiquette or rewrite cues exist, use them naturally.
- Do not use labels like "Strength:" or "Area of improvement:".
- Do not say "strong performance", "good effort", "professional standard", or other generic filler.
- Do not mention other students.
- Keep it under 90 words.
"""
        return system_prompt, user_prompt

    system_prompt = (
        "You write evidence-led evaluator remarks for long-form case-study submissions. "
        "Be concrete, selective, and avoid generic academic filler."
    )
    user_prompt = f"""
TASK:
Write one evaluator remark for a long-form case-study submission.

STUDENT:
{student_name}

SCORE:
{total_marks if total_marks is not None else "N/A"} / {max_marks if max_marks is not None else "N/A"}

PERFORMANCE BAND:
{band}

DOCUMENT SIGNALS:
paragraph_blocks={paragraph_blocks}, tables={tables}, images={images}

TOP STRENGTHS:
{json.dumps(strong_payload, indent=2)}

TOP WEAKNESSES:
{json.dumps(weak_payload, indent=2)}

INSTRUCTIONS:
- Write exactly 2 or 3 sentences.
- Mention the student name once.
- Use the strongest and weakest criteria by name.
- Explain what is convincing in the submission and what specifically limits the overall result.
- Prefer comments about framing, goals, solution design, automation choices, rationale, clarity, structure, evidence depth, and visual support when present.
- Use the supplied feedback fragments; do not invent missing details.
- Avoid generic phrases such as "strong performance", "commendable", "well-articulated", "professional standard", "further refinement", or "room for improvement".
- Do not append label-style text.
- Do not mention other students.
- Keep it under 110 words.
"""
    return system_prompt, user_prompt


def generate_llm_justified_remark(
    student_name: str,
    weak_evidence: list[dict],
    strong_evidence: list[dict] | None = None,
    total_marks: float = None,
    max_marks: float = None,
    rubric_rows: list[dict] | None = None,
    answer_text: str = "",
    doc_meta: dict[str, Any] | None = None,
) -> str:
    percentage = (total_marks / max_marks) * 100 if max_marks else None

    if percentage is not None:
        if percentage >= 80:
            band = "High"
        elif percentage >= 65:
            band = "Moderate"
        else:
            band = "Low"
    else:
        band = "Unknown"

    weak_evidence = weak_evidence[:4]
    strong_evidence = (strong_evidence or [])[:3]
    rubric_rows = rubric_rows or []

    submission_mode = _infer_submission_mode(rubric_rows, answer_text)
    strong_payload = _serialize_evidence_for_prompt(strong_evidence, limit=3)
    weak_payload = _serialize_evidence_for_prompt(weak_evidence, limit=3)

    system_prompt, user_prompt = _build_remark_prompt(
        student_name=student_name,
        submission_mode=submission_mode,
        band=band,
        total_marks=total_marks,
        max_marks=max_marks,
        strong_payload=strong_payload,
        weak_payload=weak_payload,
        doc_meta=doc_meta,
    )

    for attempt in range(2):
        try:
            openai_client = _get_openai_client()
            if openai_client is None:
                break

            response = openai_client.responses.create(
                model="gpt-4.1-mini",
                input=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": user_prompt},
                ],
                temperature=0.25,
            )

            text = _sanitize_remark((response.output_text or "").strip())
            if text:
                if student_name.lower() not in text.lower():
                    text = f"{student_name} {text[0].lower() + text[1:]}" if text else text
                return text
        except Exception:
            time.sleep(1.5)

    fallback = _build_personalized_remark_from_evidence(
        student_name=student_name,
        submission_mode=submission_mode,
        strong_evidence=strong_evidence,
        weak_evidence=weak_evidence,
        total_marks=total_marks,
        max_marks=max_marks,
    )
    return _sanitize_remark(fallback)


def _apply_scorecard_conditional_formatting(ws, total_rows, col_index):
    if total_rows <= 0:
        return

    col = get_column_letter(col_index)
    rng = f"{col}2:{col}{total_rows + 1}"

    ws.conditional_formatting.add(
        rng,
        CellIsRule("lessThan", ["40"], fill=PatternFill(start_color="FFC7CE", fill_type="solid"))
    )
    ws.conditional_formatting.add(
        rng,
        CellIsRule("between", ["40", "59.99"], fill=PatternFill(start_color="FFEB9C", fill_type="solid"))
    )
    ws.conditional_formatting.add(
        rng,
        CellIsRule("between", ["60", "79.99"], fill=PatternFill(start_color="C6EFCE", fill_type="solid"))
    )
    ws.conditional_formatting.add(
        rng,
        CellIsRule("greaterThanOrEqual", ["80"], fill=PatternFill(start_color="00B050", fill_type="solid"))
    )


def _apply_plagiarism_conditional_formatting(ws, total_rows, col_index):
    if total_rows <= 0:
        return

    col = get_column_letter(col_index)
    rng = f"{col}2:{col}{total_rows + 1}"

    ws.conditional_formatting.add(
        rng,
        CellIsRule("greaterThanOrEqual", ["80"], fill=PatternFill(start_color="FF0000", fill_type="solid"))
    )
    ws.conditional_formatting.add(
        rng,
        CellIsRule("between", ["60", "79.99"], fill=PatternFill(start_color="FFC000", fill_type="solid"))
    )


def build_master_report(
    all_students_data: List[Dict[str, Any]],
    plagiarism_data: Any
) -> str:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = os.path.join(OUTPUT_DIR, f"Final_Report_{timestamp}.xlsx")

    writer = pd.ExcelWriter(filename, engine="openpyxl")

    rubric_names = []
    for s in all_students_data:
        if s.get("results"):
            rubric_names = [r["rubric_name"] for r in s["results"]]
            break

    report_categories = []
    for s in all_students_data:
        results = s.get("results") or []
        if not results:
            continue
        rubric_rows = [{
            "Type": r.get("rubric_type", "")
        } for r in results]
        report_categories.append(_detect_rubric_category(rubric_rows))

    include_email_columns = bool(report_categories) and set(report_categories) == {"email"}

    rows = []
    for idx, s in enumerate(all_students_data, 1):
        student = s["student_name"]
        results = s.get("results") or []

        score_map = {r["rubric_name"]: float(r.get("score", 0)) for r in results}
        row = {"Sl No": idx, "Participant Name": student}
        total = 0.0

        for rn in rubric_names:
            val = score_map.get(rn, 0.0)
            row[rn] = val
            total += val

        row["Total Marks"] = total

        rubric_rows = []
        doc_meta = s.get("doc_meta", {}) or {}
        img_scores = doc_meta.get("image_scores", {}) or {}

        if img_scores and any(v > 0 for v in img_scores.values()):
            visual_score = round(
                (
                    img_scores.get("avg_relevance", 0.0)
                    + img_scores.get("avg_correctness", 0.0)
                    + img_scores.get("avg_completeness", 0.0)
                ) / 3,
                2
            )

            rubric_rows.append({
                "Rubric Name": "Visual Evidence Quality",
                "Feedback": (
                    f"Image relevance={img_scores.get('avg_relevance', 0.0):.2f}, "
                    f"correctness={img_scores.get('avg_correctness', 0.0):.2f}, "
                    f"completeness={img_scores.get('avg_completeness', 0.0):.2f}"
                ),
                "Score": visual_score,
                "Weight": 1.0,
                "Type": "visual",
            })

        for r in results:
            rubric_rows.append({
                "Rubric Name": r.get("rubric_name", ""),
                "Feedback": r.get("feedback", ""),
                "Score": float(r.get("score", 0)),
                "Weight": float(r.get("weight", 0)),
                "Type": r.get("rubric_type", ""),
            })

        weak_evidence = collect_weak_rubric_evidence(rubric_rows)
        strong_evidence = collect_strong_rubric_evidence(rubric_rows)
        max_marks = sum(float(r.get("weight", 0)) for r in results)
        rubric_category = _detect_rubric_category(rubric_rows)

        row["Remark"] = generate_llm_justified_remark(
            student_name=student,
            weak_evidence=weak_evidence,
            strong_evidence=strong_evidence,
            total_marks=row["Total Marks"],
            max_marks=max_marks,
            rubric_rows=rubric_rows,
            answer_text=s.get("answer_text", ""),
            doc_meta=doc_meta,
        )

        if include_email_columns and rubric_category == "email":
            strength, improvement = _extract_email_strength_and_improvement(
                strong_evidence=strong_evidence,
                weak_evidence=weak_evidence,
            )
            row["Strength"] = strength
            row["Area of Improvement"] = improvement

        rows.append(row)

    cols = ["Sl No", "Participant Name", *rubric_names, "Total Marks", "Remark"]
    if include_email_columns:
        cols.extend(["Strength", "Area of Improvement"])
    df_score = pd.DataFrame(rows, columns=cols)
    df_score.to_excel(writer, sheet_name="Scorecard", index=False)

    ws_score = writer.book["Scorecard"]
    total_col = cols.index("Total Marks") + 1
    _apply_scorecard_conditional_formatting(ws_score, len(rows), total_col)

    detail_rows = []
    for s in all_students_data:
        student = s["student_name"]
        for r in s.get("results") or []:
            detail_rows.append({
                "Student Name": student,
                "Rubric ID": r.get("rubric_id", ""),
                "Rubric Name": r.get("rubric_name", ""),
                "Feedback": r.get("feedback", ""),
                "Score": float(r.get("score", 0)),
                "Weight": float(r.get("weight", 0)),
                "Type": r.get("rubric_type", ""),
            })

    pd.DataFrame(detail_rows).to_excel(writer, sheet_name="Rubric_Details", index=False)

    kp_rows = []
    for s in all_students_data:
        student = s["student_name"]
        kp_map = s.get("keypoint_scores") or {}
        for rid, matches in kp_map.items():
            for m in _safe_list(matches):
                if isinstance(m, dict):
                    matched = m.get("matched_sentence", "")
                    source = "IMAGE" if matched.startswith("[IMAGE EVIDENCE]") else "TEXT"

                    kp_rows.append({
                        "Student Name": student,
                        "Rubric ID": rid,
                        "Keypoint": m.get("keypoint", ""),
                        "Similarity Score": float(m.get("similarity", 0)),
                        "Coverage": m.get("coverage", ""),
                        "Matched Sentence": matched,
                        "Evidence Source": source,
                    })

    if kp_rows:
        pd.DataFrame(kp_rows).to_excel(writer, sheet_name="Keypoints", index=False)

    pairs = []
    if isinstance(plagiarism_data, dict):
        pairs = plagiarism_data.get("pairs", [])
    elif isinstance(plagiarism_data, list):
        pairs = plagiarism_data

    plag_rows = []
    for rec in pairs:
        rec = _safe_dict(rec)
        plag_rows.append({
            "Student A": rec.get("student_a", ""),
            "Student B": rec.get("student_b", ""),
            "Similarity %": float(rec.get("score", 0)),
            "Matching Sentences": "\n".join(_safe_list(rec.get("matches"))),
            "Status": rec.get("status", ""),
        })

    df_plag = pd.DataFrame(plag_rows)
    df_plag.to_excel(writer, sheet_name="Plagiarism_Student_vs_Student", index=False)

    if not df_plag.empty:
        ws_plag = writer.book["Plagiarism_Student_vs_Student"]
        sim_col = df_plag.columns.get_loc("Similarity %") + 1
        _apply_plagiarism_conditional_formatting(ws_plag, len(df_plag), sim_col)

    ws_ev = writer.book.create_sheet("Evidence_Map")
    ws_ev.append([
        "Student",
        "Images",
        "Tables",
        "Max Table Depth",
        "Avg Image Relevance",
        "Avg Image Correctness",
        "Avg Image Completeness",
        "Visual Evidence Used",
    ])

    for s in all_students_data:
        evidence = extract_format_evidence(s.get("doc_meta", {}) or {})
        doc_meta = s.get("doc_meta", {}) or {}
        img_scores = doc_meta.get("image_scores", {}) or {}

        visual_used = (
            evidence.get("tables_detected", 0) > 0
            or evidence.get("diagrams_detected", 0) > 0
        )

        ws_ev.append([
            s["student_name"],
            evidence.get("images_detected", 0),
            evidence.get("tables_detected", 0),
            evidence.get("max_table_depth", 0.0),
            round(img_scores.get("avg_relevance", 0.0), 2),
            round(img_scores.get("avg_correctness", 0.0), 2),
            round(img_scores.get("avg_completeness", 0.0), 2),
            "YES" if visual_used else "NO",
        ])

    ws_img = writer.book.create_sheet("Image_Summaries")
    ws_img.append(["Student", "Image #", "Image Summary"])

    for s in all_students_data:
        student = s.get("student_name", "")
        doc_meta = s.get("doc_meta", {}) or {}
        images = doc_meta.get("images", []) or []

        for idx, img in enumerate(images, 1):
            ws_img.append([student, idx, img.get("summary", "")])

    writer.close()
    return filename
