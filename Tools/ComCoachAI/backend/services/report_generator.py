from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
import os
from backend.config import get_settings
from backend.utils.helpers import to_ist_str  #v2upgrades

settings = get_settings()


def generate_excel_report(test, participants):
    """Generate Excel report for a test with all participant data"""
    settings.report_dir_path.mkdir(parents=True, exist_ok=True)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Report"
    
    # ── Styles ────────────────────────────────────────────────
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ── Title Section ─────────────────────────────────────────
    ws['A1'] = f"Training Report: {test.training_name}"
    ws['A1'].font = title_font
    ws.merge_cells('A1:F1')
    
    ws['A2'] = f"Test Code: {test.test_code}"
    ws['A3'] = f"Difficulty: {test.difficulty_level}"
    ws['A4'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws['A5'] = f"Total Participants: {len(participants)}"
    
    # ── Headers ───────────────────────────────────────────────
    row = 7
    headers = ['Name', 'Email', 'Total Score (%)']
    
    # Add skill columns
    if participants and participants[0].scores:
        headers.extend(list(participants[0].scores.keys()))
    
    headers.extend(['Completed At', 'Strengths', 'Improvements'])
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # ── Data Rows ─────────────────────────────────────────────
    row += 1
    for participant in participants:
        col = 1
        
        # Basic info
        ws.cell(row=row, column=col, value=participant.name).border = border
        col += 1
        
        ws.cell(row=row, column=col, value=participant.email or "—").border = border
        col += 1
        
        ws.cell(row=row, column=col, value=round(participant.total_score, 1) if participant.total_score else 0).border = border
        col += 1
        
        # Individual skill scores
        if participant.scores:
            for skill in participants[0].scores.keys():
                score = participant.scores.get(skill, 0)
                ws.cell(row=row, column=col, value=score).border = border
                col += 1
        
        # Timestamp
        completed = to_ist_str(participant.completed_at) if participant.completed_at else "—"  #v2upgrades
        ws.cell(row=row, column=col, value=completed).border = border
        col += 1
        
        # Strengths
        ws.cell(row=row, column=col, value=participant.strengths or "—").border = border
        col += 1
        
        # Improvements
        ws.cell(row=row, column=col, value=participant.improvements or "—").border = border
        
        row += 1
    
    # ── Average Row ───────────────────────────────────────────
    row += 1
    ws.cell(row=row, column=1, value="AVERAGE").font = Font(bold=True)
    ws.cell(row=row, column=1).border = border
    
    # Skip email column
    ws.cell(row=row, column=2, value="—").border = border
    
    # Average total score
    avg_total = sum(p.total_score for p in participants if p.total_score) / len(participants)
    ws.cell(row=row, column=3, value=round(avg_total, 1)).font = Font(bold=True)
    ws.cell(row=row, column=3).border = border
    
    # Average skill scores
    if participants and participants[0].scores:
        col = 4
        for skill in participants[0].scores.keys():
            scores = [p.scores.get(skill, 0) for p in participants if p.scores and skill in p.scores]
            avg = sum(scores) / len(scores) if scores else 0
            ws.cell(row=row, column=col, value=round(avg, 1)).font = Font(bold=True)
            ws.cell(row=row, column=col).border = border
            col += 1
    
    # ── Column Widths ─────────────────────────────────────────
    ws.column_dimensions['A'].width = 20  # Name
    ws.column_dimensions['B'].width = 25  # Email
    ws.column_dimensions['C'].width = 15  # Total Score
    
    # Skill columns
    if participants and participants[0].scores:
        for idx in range(len(participants[0].scores)):
            col_letter = chr(68 + idx)  # D, E, F, etc.
            ws.column_dimensions[col_letter].width = 12
    
    # Timestamp, Strengths, Improvements
    last_cols = ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']
    for col_letter in last_cols[-3:]:
        if col_letter in ws.column_dimensions:
            ws.column_dimensions[col_letter].width = 40
    
    # ── Save ──────────────────────────────────────────────────
    filename = f"{test.test_code}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = settings.report_dir_path / filename
    
    wb.save(filepath)
    return str(filepath)
